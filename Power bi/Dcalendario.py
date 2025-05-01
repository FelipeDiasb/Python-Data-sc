import pandas as pd
import calendar
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font  # Adicionando Font aqui
from datetime import datetime

# Carregar a planilha de feriados
feriados_df = pd.read_excel('Power bi/Feriados/feriados_nacionais.xls')

# Garantir que a coluna 'Data' esteja no formato datetime
feriados_df['Data'] = pd.to_datetime(feriados_df['Data'], errors='coerce')

# Verificar se há valores inválidos (NaT) após a conversão
feriados_df = feriados_df.dropna(subset=['Data'])  # Remover valores inválidos

# Converter as datas para o formato dd/mm
feriados_data = feriados_df['Data'].dt.strftime('%d/%m').tolist()

# Criar uma planilha Excel
wb = Workbook()
ws = wb.active
ws.title = "Calendário 2025"

# Títulos das colunas (dias da semana)
dias_da_semana = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
for col, dia in enumerate(dias_da_semana, start=1):
    ws.cell(row=1, column=col, value=dia).alignment = Alignment(horizontal='center', vertical='center')

# Função para preencher um mês no Excel
def preencher_mes(ano, mes, linha_inicial=2):
    mes_calendario = calendar.monthcalendar(ano, mes)  # Lista de semanas do mês
    for semana in mes_calendario:
        for col, dia in enumerate(semana, start=1):
            if dia != 0:  # Ignorar os zeros (dias em branco)
                # Formatar a célula
                data_dia = f"{dia:02d}/{mes:02d}"
                cell = ws.cell(row=linha_inicial, column=col, value=dia)
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Verificar se é feriado
                if data_dia in feriados_data:
                    # Colorir a célula de feriado
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")  # Aplicando fonte em negrito e branca

        linha_inicial += 1

# Gerar o calendário para o ano de 2025
ano = 2025
linha_atual = 2  # Começar a preencher os meses a partir da linha 2
for mes in range(1, 13):  # Para todos os meses de 1 a 12
    # Definir o título do mês
    ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=7)
    mes_nome = calendar.month_name[mes]
    ws.cell(row=linha_atual, column=1, value=mes_nome).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=linha_atual, column=1).font = Font(bold=True)
    
    linha_atual += 1
    preencher_mes(ano, mes, linha_atual)
    linha_atual += 7  # Pular para a próxima linha para o próximo mês

# Salvar o arquivo
wb.save("Calendário_2025_com_feriados.xlsx")

print("Calendário 2025 com feriados extraídos da planilha gerado com sucesso!")
